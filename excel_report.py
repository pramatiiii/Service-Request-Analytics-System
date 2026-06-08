"""
reports/excel_report.py
========================
Generates a fully formatted Excel workbook: output/analytics_report.xlsx

Library used: openpyxl
-----------------------
openpyxl lets us create and style Excel files programmatically.
Unlike pandas.to_excel() which only writes raw data, openpyxl gives
full control over colours, fonts, column widths, frozen rows, and more.

Key hierarchy:
    Workbook   → the entire .xlsx file
    Worksheet  → one tab/sheet inside the workbook
    Cell       → one individual cell on a sheet

Install: pip install openpyxl
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from analytics.filters import get_high_priority_requests, get_completed_requests
from analytics.stats   import (
    status_breakdown,
    priority_breakdown,
    average_resolution_time,
    top_failure_categories,
    technician_performance,
)


# ── Colour constants ───────────────────────────────────────────────────────────
# Note: openpyxl requires hex WITHOUT the # prefix (e.g. "A31D1D" not "#A31D1D")
# These match the dark colour palette used throughout the project.
COL_HEADER_RED    = "A31D1D"
COL_HEADER_BROWN  = "754E1A"
COL_HEADER_GREEN  = "254F22"
COL_HEADER_MAROON = "810B38"
COL_HEADER_OLIVE  = "543A14"
COL_HEADER_DARK   = "313E17"
COL_ROW_DARK      = "141414"
COL_ROW_ALT       = "1e1e1e"
COL_WHITE_TEXT    = "EEEEEE"
COL_WHITE         = "FFFFFF"


# ── Reusable style helper functions ───────────────────────────────────────────

def style_header(cell, bg_color=COL_HEADER_RED):
    """
    Apply header styling to a cell: bold white text on a coloured background.

    Parameters
    ----------
    cell     : openpyxl Cell object
    bg_color : hex colour string (without #) for the background fill

    This is a helper function — it exists purely to avoid repeating the same
    5 lines of styling code for every header cell across all 6 sheets.
    """
    cell.font      = Font(bold=True, color=COL_WHITE, size=11)
    cell.fill      = PatternFill("solid", start_color=bg_color)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border    = Border(bottom=Side(style="thin", color=COL_WHITE))


def style_data(cell, bg_color=COL_ROW_DARK):
    """
    Apply data row styling to a cell: light text on a dark background.

    Parameters
    ----------
    cell     : openpyxl Cell object
    bg_color : hex colour for the row background (alternates for zebra striping)
    """
    cell.font      = Font(color=COL_WHITE_TEXT, size=10)
    cell.fill      = PatternFill("solid", start_color=bg_color)
    cell.alignment = Alignment(horizontal="left", vertical="center")


def auto_fit_columns(worksheet):
    """
    Automatically set each column's width based on the longest content in that column.

    Why is this needed?
    Excel defaults to a narrow column width that cuts off long values.
    We calculate the maximum character length in each column and set the
    width accordingly — the same as double-clicking a column border in Excel.

    get_column_letter(n) converts a column number to its letter:
        1 → "A", 2 → "B", 3 → "C" ... 26 → "Z", 27 → "AA"
    This is required because Excel column widths are indexed by letter, not number.

    min(..., 40) caps width at 40 to prevent extremely wide columns.
    """
    for column in worksheet.columns:
        max_length = max(
            (len(str(cell.value or "")) for cell in column),
            default=10
        )
        col_letter = get_column_letter(column[0].column)
        worksheet.column_dimensions[col_letter].width = min(max_length + 4, 40)


def row_color(row_index):
    """
    Return alternating row background colours for zebra striping.

    Zebra striping (alternating light/dark rows) makes long tables
    much easier to read by giving the eye a visual anchor per row.

    The modulo operator (%) gives the remainder of division.
    row_index % 2 == 0 → True for even rows, False for odd rows.
    """
    return COL_ROW_ALT if row_index % 2 == 0 else COL_ROW_DARK


# ── Column definitions ─────────────────────────────────────────────────────────
# The 10 column headers used in any sheet that shows full request data.
REQUEST_HEADERS = [
    "SR_ID", "Asset_ID", "Customer_Name", "Machine_Type",
    "Reported_Date", "Priority", "Status", "Technician",
    "Resolution_Time_Hours", "Failure_Category",
]

# Colours to highlight Priority and Status cells specifically
PRIORITY_CELL_COLORS = {"High": "A31D1D", "Medium": "754E1A", "Low": "254F22"}
STATUS_CELL_COLORS   = {"Open": "810B38", "In Progress": "543A14", "Completed": "313E17"}


def _write_request_headers(worksheet, header_color):
    """Write the 10 request column headers to row 1 of a worksheet."""
    for col_idx, header in enumerate(REQUEST_HEADERS, start=1):
        style_header(worksheet.cell(row=1, column=col_idx, value=header), header_color)


def _write_request_rows(worksheet, requests):
    """
    Write ServiceRequest objects as data rows starting at row 2.

    Priority and Status columns get their own specific colours so they
    stand out visually in the table — an extra level of readability.
    """
    for row_idx, sr in enumerate(requests, start=2):
        bg = row_color(row_idx)

        # Build the row values in the same order as REQUEST_HEADERS
        values = [
            sr.sr_id, sr.asset_id, sr.customer_name, sr.machine_type,
            sr.reported_date, sr.priority, sr.status, sr.technician,
            sr.resolution_time_hours, sr.failure_category,
        ]

        for col_idx, value in enumerate(values, start=1):
            cell = worksheet.cell(row=row_idx, column=col_idx, value=value)

            if col_idx == 6:   # Priority column — use priority-specific colour
                style_data(cell, PRIORITY_CELL_COLORS.get(str(value), bg))
            elif col_idx == 7: # Status column — use status-specific colour
                style_data(cell, STATUS_CELL_COLORS.get(str(value), bg))
            else:
                style_data(cell, bg)


# ── Main export function ───────────────────────────────────────────────────────

def generate_excel_report(requests, df, output_dir="output"):
    """
    Create a fully formatted Excel workbook with 6 sheets.

    Parameters
    ----------
    requests   : list of ServiceRequest objects
    df         : pandas DataFrame (used for tech performance and averages)
    output_dir : folder to save the file

    Returns
    -------
    str — file path of the saved workbook

    Sheets created:
        1. Summary               — key metrics at a glance
        2. All Requests          — every record with colour-coded Priority/Status
        3. High Priority         — filtered view: High priority only
        4. Completed Requests    — filtered view: Completed only
        5. Technician Performance — grouped stats per technician
        6. Failure Categories    — ranked failure types
    """
    os.makedirs(output_dir, exist_ok=True)
    filepath = os.path.join(output_dir, "analytics_report.xlsx")

    wb = Workbook()  # Create a new empty workbook in memory

    # ── Sheet 1: Summary ──────────────────────────────────────────────────────
    ws_summary = wb.active          # The first sheet already exists; grab it
    ws_summary.title = "Summary"
    ws_summary.sheet_properties.tabColor = COL_HEADER_RED

    # Merge cells A1:C1 to create a wide title banner
    ws_summary.merge_cells("A1:C1")
    title_cell = ws_summary["A1"]
    title_cell.value     = "SERVICE REQUEST ANALYTICS REPORT"
    title_cell.font      = Font(bold=True, size=14, color=COL_WHITE)
    title_cell.fill      = PatternFill("solid", start_color=COL_HEADER_RED)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_summary.row_dimensions[1].height = 30   # Taller title row

    # Gather stats
    sb       = status_breakdown(requests)
    pb       = priority_breakdown(requests)
    avg_time = average_resolution_time(df)

    # Each tuple is (label, value) — written as one row
    summary_rows = [
        ("Metric",              "Value"),
        ("Total Requests",      len(requests)),
        ("Open",                sb.get("Open", 0)),
        ("In Progress",         sb.get("In Progress", 0)),
        ("Completed",           sb.get("Completed", 0)),
        ("High Priority",       pb.get("High", 0)),
        ("Medium Priority",     pb.get("Medium", 0)),
        ("Low Priority",        pb.get("Low", 0)),
        ("Avg Resolution (hrs)", avg_time),
    ]

    for row_idx, (label, value) in enumerate(summary_rows, start=2):
        bg = row_color(row_idx)
        label_cell = ws_summary.cell(row=row_idx, column=1, value=label)
        value_cell = ws_summary.cell(row=row_idx, column=2, value=value)

        if row_idx == 2:   # Sub-header row gets its own styling
            style_header(label_cell, COL_HEADER_BROWN)
            style_header(value_cell, COL_HEADER_BROWN)
        else:
            style_data(label_cell, bg)
            style_data(value_cell, bg)

    auto_fit_columns(ws_summary)

    # ── Sheet 2: All Requests ─────────────────────────────────────────────────
    ws_all = wb.create_sheet("All Requests")
    ws_all.sheet_properties.tabColor = COL_HEADER_OLIVE
    _write_request_headers(ws_all, COL_HEADER_OLIVE)
    _write_request_rows(ws_all, requests)
    ws_all.freeze_panes = "A2"   # Keep header row visible when scrolling
    auto_fit_columns(ws_all)

    # ── Sheet 3: High Priority ────────────────────────────────────────────────
    ws_high = wb.create_sheet("High Priority")
    ws_high.sheet_properties.tabColor = COL_HEADER_RED
    _write_request_headers(ws_high, COL_HEADER_RED)
    _write_request_rows(ws_high, get_high_priority_requests(requests))
    ws_high.freeze_panes = "A2"
    auto_fit_columns(ws_high)

    # ── Sheet 4: Completed Requests ───────────────────────────────────────────
    ws_completed = wb.create_sheet("Completed Requests")
    ws_completed.sheet_properties.tabColor = COL_HEADER_GREEN
    _write_request_headers(ws_completed, COL_HEADER_GREEN)
    _write_request_rows(ws_completed, get_completed_requests(requests))
    ws_completed.freeze_panes = "A2"
    auto_fit_columns(ws_completed)

    # ── Sheet 5: Technician Performance ──────────────────────────────────────
    ws_tech = wb.create_sheet("Technician Performance")
    ws_tech.sheet_properties.tabColor = COL_HEADER_DARK

    tech_headers = ["Technician", "Completed Requests", "Avg Resolution (hrs)"]
    for col_idx, header in enumerate(tech_headers, start=1):
        style_header(ws_tech.cell(row=1, column=col_idx, value=header), COL_HEADER_DARK)

    perf_df = technician_performance(df)
    for row_idx, row in enumerate(perf_df.itertuples(index=False), start=2):
        bg = row_color(row_idx)
        style_data(ws_tech.cell(row=row_idx, column=1, value=row.Technician), bg)
        style_data(ws_tech.cell(row=row_idx, column=2, value=row.Completed_Requests), bg)
        style_data(ws_tech.cell(row=row_idx, column=3, value=row.Avg_Resolution_Time), bg)

    auto_fit_columns(ws_tech)

    # ── Sheet 6: Failure Categories ───────────────────────────────────────────
    ws_fail = wb.create_sheet("Failure Categories")
    ws_fail.sheet_properties.tabColor = COL_HEADER_MAROON

    for col_idx, header in enumerate(["Rank", "Failure Category", "Occurrences"], start=1):
        style_header(ws_fail.cell(row=1, column=col_idx, value=header), COL_HEADER_MAROON)

    for rank, (category, count) in enumerate(top_failure_categories(requests, n=8), start=1):
        bg = row_color(rank)
        style_data(ws_fail.cell(row=rank + 1, column=1, value=rank), bg)
        style_data(ws_fail.cell(row=rank + 1, column=2, value=category), bg)
        style_data(ws_fail.cell(row=rank + 1, column=3, value=count), bg)

    auto_fit_columns(ws_fail)

    # ── Save workbook to disk ─────────────────────────────────────────────────
    wb.save(filepath)
    print(f"[✓] Excel report saved  → {filepath}")
    return filepath
